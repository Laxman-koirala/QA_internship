import openpyxl


def getTotalrows(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet =  workbook.get_sheet_by_name(sheetName)
    return (sheet.max_row)

def getTotalcolumns(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet =  workbook.get_sheet_by_name(sheetName)
    return(sheet.max_column)

def readData(file,sheetName,rownum,columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet =  workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum,column=columnnum).value

def writeData(file1,sheetName,rownum,columnnum,data,file2):
    workbook = openpyxl.load_workbook(file1)
    sheet =  workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum,column=columnnum).value = data
    workbook.save(file2)


